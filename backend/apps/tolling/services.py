"""Service voor CSV-import en export van tolheffing-events."""
from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from xml.sax.saxutils import escape

from django.db import IntegrityError, transaction
from django.utils import timezone

from apps.fleet.bedrijven import bouw_bedrijf_index, zoek_bedrijf_in_index
from apps.fleet.ritnummers import bouw_ritnummer_index, zoek_in_index

from .models import PrivateTollRegistration, TollingEvent, TollingImportBatch, normalize_plate

logger = logging.getLogger(__name__)

REQUIRED_HEADERS = {'start date', 'end date', 'distance', 'amount', 'license plate'}


@dataclass
class ImportResult:
    batch: TollingImportBatch
    imported: int
    duplicates: int
    invalid: int
    total: int


def _parse_decimal(value: str) -> Decimal | None:
    """Parse een bedrag of afstand uit de CSV.

    Ondersteunt zowel Europese notatie (1.234,56) als Amerikaanse (1,234.56)
    en negeert valutatekens, (non-breaking) spaties en andere opmaak. Zonder
    deze normalisatie werden zulke regels stilzwijgend als 'ongeldig'
    overgeslagen.
    """
    if value is None:
        return None
    s = str(value).replace('\u00a0', ' ').strip()
    if not s:
        return None
    # Alles behalve cijfers, scheidingstekens en teken verwijderen (€, EUR, spaties, ...).
    s = re.sub(r'[^0-9,.+-]', '', s)
    if not s or s in ('-', '+', '.', ','):
        return None
    if _looks_european(s):
        s = s.replace('.', '').replace(',', '.')
    else:
        # Komma kan hier alleen duizendtalscheiding zijn.
        s = s.replace(',', '')
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _looks_european(value: str) -> bool:
    """Detect European decimal notation: uses comma as decimal separator."""
    s = str(value).strip()
    if ',' in s and s.rfind(',') > s.rfind('.'):
        return True
    return False


def _parse_datetime(value: str) -> datetime | None:
    """Parse een timestamp uit de CSV.

    Regel: als de string GEEN expliciete tijdzone-info bevat (bijv. 'Z',
    '+02:00', '-05:00') wordt de tijd geïnterpreteerd als **lokale tijd**
    (`settings.TIME_ZONE`, doorgaans Europe/Amsterdam voor deze deploy).
    Alleen wanneer de string zelf tijdzone-info draagt wordt die
    gerespecteerd en omgezet naar UTC bij opslag.

    Achtergrond: tolheffing-CSV's (DKV/AS24/etc.) gebruiken standaard
    lokale tijd zonder tz-suffix. Python 3.11's `datetime.fromisoformat`
    accepteert óók `'YYYY-MM-DD HH:MM:SS'`; dat leverde eerder een stille
    UTC-interpretatie op → duplicate imports met 2-uur-shift. Nu is de
    default consistent: geen tz-info = lokaal.
    """
    if not value:
        return None
    v = value.strip()
    dt: datetime | None = None
    # 1) ISO 8601 (met of zonder Z)
    try:
        if v.endswith('Z'):
            v_iso = v[:-1] + '+00:00'
        else:
            v_iso = v
        dt = datetime.fromisoformat(v_iso)
    except ValueError:
        dt = None
    # 2) Fallback-formaten (Europees / Amerikaans)
    if dt is None:
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S',
                    '%d-%m-%Y %H:%M:%S', '%d/%m/%Y %H:%M:%S'):
            try:
                dt = datetime.strptime(v, fmt)
                break
            except ValueError:
                continue
    if dt is None:
        return None
    # 3) Als er geen tijdzone bij de string zat → interpreteer als lokaal.
    if dt.tzinfo is None:
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return dt


def _sniff_reader(raw: bytes) -> csv.DictReader:
    text = raw.decode('utf-8-sig', errors='replace')
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
    except csv.Error:
        class _D(csv.excel):
            delimiter = ','
        dialect = _D
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    return reader


def _header_map(fieldnames: list[str]) -> dict[str, str]:
    """Map arbitrary case/whitespace headers to canonical keys."""
    mapping = {}
    for name in fieldnames or []:
        key = name.strip().lower()
        mapping[key] = name
    return mapping


def build_vehicle_lookup() -> dict:
    """Genormaliseerd kenteken -> Vehicle.

    Wordt gebruikt om bij de import het ritnummer, de wagen en het bedrijf
    vast te leggen. Staat hetzelfde kenteken meerdere keren in de vloot, dan
    wint het actieve voertuig (die worden als laatste ingelezen).
    """
    from apps.fleet.models import Vehicle

    lookup: dict = {}
    for v in Vehicle.objects.select_related('bedrijf').order_by('actief', 'created_at'):
        key = normalize_plate(v.kenteken)
        if key:
            lookup[key] = v
    return lookup


def import_csv(file_obj, user, filename: str = '') -> ImportResult:
    """Import a tolling CSV file. Idempotent — duplicates are skipped."""
    raw = file_obj.read()
    if isinstance(raw, str):
        raw = raw.encode('utf-8')
    reader = _sniff_reader(raw)
    header_map = _header_map(reader.fieldnames or [])

    missing = REQUIRED_HEADERS - set(header_map.keys())
    batch = TollingImportBatch.objects.create(
        uploaded_by=user if getattr(user, 'is_authenticated', False) else None,
        filename=filename or getattr(file_obj, 'name', '') or '',
    )
    if missing:
        batch.error_message = f"Ontbrekende kolommen: {', '.join(sorted(missing))}"
        batch.save(update_fields=['error_message'])
        return ImportResult(batch=batch, imported=0, duplicates=0, invalid=0, total=0)

    def col(row, key):
        return row.get(header_map[key], '') if key in header_map else ''

    imported = duplicates = invalid = total = 0
    new_events: list[TollingEvent] = []
    # Momentopname van de vloot: het ritnummer en het bedrijf die op de datum
    # van de passage golden worden op het event vastgelegd, zodat een latere
    # wijziging de historie niet met terugwerkende kracht herschrijft en een
    # bestand dat over een wissel heen loopt toch goed verdeeld wordt.
    vehicle_lookup = build_vehicle_lookup()
    wagen_ids = {v.pk for v in vehicle_lookup.values() if v is not None}
    ritnummer_index = bouw_ritnummer_index(wagen_ids)
    bedrijf_index = bouw_bedrijf_index(wagen_ids)
    tijdzone = timezone.get_current_timezone()

    for row in reader:
        total += 1
        start = _parse_datetime(col(row, 'start date'))
        end = _parse_datetime(col(row, 'end date'))
        distance = _parse_decimal(col(row, 'distance'))
        amount = _parse_decimal(col(row, 'amount'))
        plate_raw = str(col(row, 'license plate') or '').strip()
        obu = str((col(row, 'obu') if 'obu' in header_map else '') or '').strip()

        if not (start and end and distance is not None and amount is not None and plate_raw):
            invalid += 1
            continue

        plate_norm = normalize_plate(plate_raw)
        vehicle = vehicle_lookup.get(plate_norm)

        # Datum van de passage in de lokale tijdzone; daarop wordt het
        # ritnummer opgezocht.
        passagedatum = (
            timezone.localtime(start, tijdzone).date()
            if timezone.is_aware(start) else start.date()
        )

        new_events.append(TollingEvent(
            batch=batch,
            start_at=start,
            end_at=end,
            distance_km=distance,
            amount=amount,
            license_plate_raw=plate_raw,
            license_plate_normalized=plate_norm,
            obu=obu,
            vehicle=vehicle,
            ritnummer=zoek_in_index(ritnummer_index, vehicle, passagedatum),
            bedrijf_id=zoek_bedrijf_in_index(bedrijf_index, vehicle, passagedatum),
        ))

    # Bulk insert; on unique conflict fall back to per-row insert to count duplicates.
    if new_events:
        try:
            with transaction.atomic():
                TollingEvent.objects.bulk_create(new_events)
                imported = len(new_events)
        except IntegrityError:
            imported = 0
            for ev in new_events:
                try:
                    with transaction.atomic():
                        ev.save()
                        imported += 1
                except IntegrityError:
                    duplicates += 1

    batch.rows_total = total
    batch.rows_imported = imported
    batch.rows_duplicate = duplicates
    batch.rows_invalid = invalid
    batch.save(update_fields=['rows_total', 'rows_imported', 'rows_duplicate', 'rows_invalid'])

    # Match freshly imported events against existing private registrations
    try:
        events_qs = TollingEvent.objects.filter(batch=batch)
        match_events_to_private_registrations(events_qs)
    except Exception as exc:  # pragma: no cover
        logger.warning("Private-tolregistratie matching mislukt voor batch %s: %s", batch.id, exc)

    return ImportResult(batch=batch, imported=imported, duplicates=duplicates, invalid=invalid, total=total)


# ---------- Private toll matching ----------

def match_events_to_private_registrations(events_qs) -> int:
    """Voor elke event in `events_qs` zoek een matchende PrivateTollRegistration en markeer als privé.

    Match-criterium: zelfde genormaliseerde kenteken + het event-interval [start_at, end_at]
    overlapt met het datum-venster [begin_tijd, eind_tijd] van de registratie (lokaal).
    Bestaande privé-links worden niet overschreven, en events die al aan een factuurregel
    gekoppeld zijn worden overgeslagen.
    Retourneert aantal nieuw als privé gemarkeerde events.
    """
    tz = timezone.get_current_timezone()
    count = 0
    for ev in events_qs.filter(is_private=False, invoice_line__isnull=True).iterator():
        local_start = ev.start_at.astimezone(tz) if timezone.is_aware(ev.start_at) else ev.start_at
        local_end = ev.end_at.astimezone(tz) if (ev.end_at and timezone.is_aware(ev.end_at)) else ev.end_at
        # Zoek registraties op zelfde kenteken die op de start-datum OF einde-datum vallen
        # en waarvan het tijdvenster overlapt met het event-interval.
        candidates = PrivateTollRegistration.objects.filter(
            license_plate_normalized=ev.license_plate_normalized,
            datum__in={local_start.date(), local_end.date() if local_end else local_start.date()},
        )
        reg = None
        for cand in candidates:
            reg_start = timezone.make_aware(datetime.combine(cand.datum, cand.begin_tijd), tz)
            reg_end = timezone.make_aware(datetime.combine(cand.datum, cand.eind_tijd), tz)
            ev_end = ev.end_at if ev.end_at else ev.start_at
            if ev.start_at < reg_end and ev_end > reg_start:
                reg = cand
                break
        if reg is not None:
            ev.is_private = True
            ev.private_registration = reg
            ev.save(update_fields=['is_private', 'private_registration'])
            count += 1
    return count


def match_private_registration_to_events(reg: PrivateTollRegistration) -> int:
    """Markeer alle bestaande, niet-gefactureerde events die overlappen met `reg` als privé.

    Overlap: event.start_at < reg_end AND event.end_at > reg_start.
    Retourneert aantal nieuw als privé gemarkeerde events.
    """
    tz = timezone.get_current_timezone()
    start_dt = timezone.make_aware(datetime.combine(reg.datum, reg.begin_tijd), tz)
    end_dt = timezone.make_aware(datetime.combine(reg.datum, reg.eind_tijd), tz)
    if end_dt <= start_dt:
        return 0

    qs = TollingEvent.objects.filter(
        license_plate_normalized=reg.license_plate_normalized,
        start_at__lt=end_dt,
        end_at__gt=start_dt,
        invoice_line__isnull=True,
    ).exclude(is_private=True)
    updated = qs.update(is_private=True, private_registration=reg)
    return updated


def unmatch_private_registration(reg: PrivateTollRegistration) -> int:
    """Verbreek alle links van events naar deze registratie (bij verwijderen/wijzigen)."""
    qs = TollingEvent.objects.filter(private_registration=reg)
    return qs.update(is_private=False, private_registration=None)


# ---------- Exports ----------

def _xlsx_tekst(waarde) -> str:
    """Maak tekst veilig voor een Excel-cel.

    Excel voert een cel die met =, +, - of @ begint uit als formule. Kentekens
    en ritnummers komen deels uit geimporteerde bestanden, dus die tekst wordt
    met een apostrof onschadelijk gemaakt.
    """
    tekst = str(waarde or '')
    if tekst[:1] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + tekst
    return tekst


def _pdf_tekst(waarde) -> str:
    """Maak tekst veilig voor een reportlab Paragraph, die opmaaktags leest."""
    return escape(str(waarde or ''))


def export_events_xlsx(events, plate_label: str, period_label: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font

    wb = Workbook()
    ws = wb.active
    ws.title = 'Tolheffing'
    ws.append([f'Kenteken: {plate_label}'])
    ws.append([f'Periode: {period_label}'])
    ws.append([])
    ws.append(['Startdatum', 'Einddatum', 'Ritnummer', 'Type', 'Afstand (km)', 'Bedrag (€)', 'Gefactureerd'])
    weekend_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    private_fill = PatternFill(start_color='EDE9FE', end_color='EDE9FE', fill_type='solid')
    private_font = Font(color='5B21B6')
    total_km = Decimal('0')
    total_amount = Decimal('0')
    private_km = Decimal('0')
    private_amount = Decimal('0')
    weekend_km = Decimal('0')
    weekend_amount = Decimal('0')
    weekday_km = Decimal('0')
    weekday_amount = Decimal('0')
    # Subtotalen per ritnummer, zodat een wagen die in de periode van rit is
    # gewisseld beide ritnummers apart terugkomt in de uitdraai.
    per_rit: dict[str, list] = {}
    for e in events:
        is_priv = bool(getattr(e, 'is_private', False))
        is_wknd = bool(e.start_at and e.start_at.isoweekday() >= 6)
        rit = (getattr(e, 'ritnummer', '') or '').strip()
        if is_priv:
            type_label = 'Privé'
        elif is_wknd:
            type_label = 'Weekend'
        else:
            type_label = 'Doordeweeks'
        ws.append([
            e.start_at.strftime('%Y-%m-%d %H:%M'),
            e.end_at.strftime('%Y-%m-%d %H:%M'),
            _xlsx_tekst(rit),
            type_label,
            float(e.distance_km),
            float(e.amount),
            'Ja' if e.invoiced_at else ('Privé' if is_priv else 'Nee'),
        ])
        row_idx = ws.max_row
        if is_priv:
            for col in range(1, 8):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = private_fill
                cell.font = private_font
            private_km += Decimal(e.distance_km)
            private_amount += Decimal(e.amount)
        else:
            if is_wknd:
                for col in range(1, 8):
                    ws.cell(row=row_idx, column=col).fill = weekend_fill
                weekend_km += Decimal(e.distance_km)
                weekend_amount += Decimal(e.amount)
            else:
                weekday_km += Decimal(e.distance_km)
                weekday_amount += Decimal(e.amount)
            total_km += Decimal(e.distance_km)
            total_amount += Decimal(e.amount)
            bucket = per_rit.setdefault(rit, [Decimal('0'), Decimal('0'), 0])
            bucket[0] += Decimal(e.distance_km)
            bucket[1] += Decimal(e.amount)
            bucket[2] += 1
    ws.append([])
    ws.append(['', 'Doordeweeks', '', '', float(weekday_km), float(weekday_amount), ''])
    ws.append(['', 'Weekend', '', '', float(weekend_km), float(weekend_amount), ''])
    if private_km > 0 or private_amount > 0:
        ws.append(['', 'Privé (niet doorbelast)', '', '', float(private_km), float(private_amount), ''])
        row_idx = ws.max_row
        for col in range(1, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = private_fill
            cell.font = private_font
    ws.append(['', 'Totaal tolkosten', '', '', float(total_km), float(total_amount), ''])
    if len(per_rit) > 1:
        ws.append([])
        ws.append(['Per ritnummer', '', '', '', '', '', ''])
        for rit in sorted(per_rit):
            km, bedrag, aantal = per_rit[rit]
            ws.append([
                '', _xlsx_tekst(rit) or 'Geen ritnummer', f'{aantal} regels', '',
                float(km), float(bedrag), '',
            ])
    for col_idx, width in enumerate([20, 20, 16, 14, 15, 15, 15], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# Dachser-export
# ---------------------------------------------------------------------------

DACHSER_HEADER_FILL = '002060'   # donkerblauw
DACHSER_HEADER_FONT = 'FFC000'   # goud/geel
DACHSER_LOGO_BLUE = (12, 77, 162)

DACHSER_COLUMNS = [
    ('Route', 'route', 12),
    ('Carrier', 'carrier', 16),
    ('Country', 'country', 12),
    ('License plate', 'license_plate', 18),
    ('Total tol kilometers', 'total_km', 24),
    ('Amount EUR.', 'amount', 16),
    ('Date', 'date', 13),
]


def _dachser_logo_bytes() -> bytes | None:
    """Lever de logo-afbeelding voor bovenaan de Dachser-export.

    Als `apps/tolling/assets/dachser_logo.png` bestaat wordt die gebruikt.
    Anders wordt het woordmerk met Pillow getekend, zodat de export ook
    zonder los logobestand een afbeelding bovenaan heeft.
    """
    import os

    custom = os.path.join(os.path.dirname(__file__), 'assets', 'dachser_logo.png')
    if os.path.exists(custom):
        try:
            with open(custom, 'rb') as fh:
                return fh.read()
        except OSError:
            logger.warning('Kon dachser_logo.png niet lezen, val terug op gegenereerd logo')

    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        return None

    width, height = 640, 170
    img = Image.new('RGB', (width, height), 'white')
    draw = ImageDraw.Draw(img)

    def _font(size: int):
        for path in (
            '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
            '/usr/local/lib/python3.12/site-packages/reportlab/fonts/VeraBd.ttf',
        ):
            if os.path.exists(path):
                try:
                    return ImageFont.truetype(path, size)
                except OSError:
                    continue
        return ImageFont.load_default()

    draw.text((6, 4), 'DACHSER', font=_font(84), fill=DACHSER_LOGO_BLUE)
    draw.text((8, 104), 'Intelligent Logistics', font=_font(44), fill=(20, 20, 20))

    buf = io.BytesIO()
    img.save(buf, format='PNG')
    return buf.getvalue()


def build_dachser_export_xlsx(rows: list[dict]) -> bytes:
    """Bouw het Dachser-exportbestand.

    `rows` bevat dicts met de keys: route, carrier, country, license_plate,
    total_km (float), amount (float) en date (datetime.date).
    Opmaak: logo bovenaan, donkerblauwe koprij met gouden vette tekst,
    randen rond alle cellen en een autofilter per kolom.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Toll'

    logo_rows = 5
    header_row = logo_rows + 1

    logo = _dachser_logo_bytes()
    if logo:
        try:
            from openpyxl.drawing.image import Image as XlsxImage

            img = XlsxImage(io.BytesIO(logo))
            # ~5 rijen hoog houden, breedte proportioneel meeschalen
            target_height = 96
            if img.height:
                ratio = target_height / img.height
                img.width = int(img.width * ratio)
                img.height = target_height
            img.anchor = 'A1'
            ws.add_image(img)
        except Exception:  # pragma: no cover - logo mag de export nooit blokkeren
            logger.exception('Kon logo niet invoegen in Dachser-export')

    for r in range(1, logo_rows + 1):
        ws.row_dimensions[r].height = 20

    header_fill = PatternFill(start_color=DACHSER_HEADER_FILL,
                              end_color=DACHSER_HEADER_FILL, fill_type='solid')
    header_font = Font(bold=True, color=DACHSER_HEADER_FONT, size=11)
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Donkerblauwe balk boven de koprij (zoals in het voorbeeldbestand)
    for col_idx in range(1, len(DACHSER_COLUMNS) + 1):
        ws.cell(row=logo_rows, column=col_idx).fill = header_fill

    for col_idx, (label, _key, width) in enumerate(DACHSER_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 20

    for offset, row in enumerate(rows):
        excel_row = header_row + 1 + offset
        values = [
            # Vrije tekst wordt onschadelijk gemaakt: Excel voert een cel die
            # met = + - of @ begint anders uit als formule.
            _xlsx_tekst(row.get('route')),
            _xlsx_tekst(row.get('carrier')),
            _xlsx_tekst(row.get('country') or 'NL'),
            _xlsx_tekst(row.get('license_plate')),
            float(row.get('total_km') or 0),
            float(row.get('amount') or 0),
            row.get('date'),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = border
            if col_idx == 5:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            elif col_idx == 6:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            elif col_idx == 7:
                cell.number_format = 'dd-mm-yyyy'
                cell.alignment = Alignment(horizontal='left')

    last_row = header_row + max(len(rows), 1)
    last_col = get_column_letter(len(DACHSER_COLUMNS))
    ws.auto_filter.ref = f'A{header_row}:{last_col}{last_row}'
    ws.freeze_panes = f'A{header_row + 1}'

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_events_pdf(events, plate_label: str, period_label: str) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    styles = getSampleStyleSheet()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
    )
    elems = [
        Paragraph(f'Tolheffing overzicht — kenteken {_pdf_tekst(plate_label)}', styles['Title']),
        Paragraph(f'Periode: {_pdf_tekst(period_label)}', styles['Normal']),
        Spacer(1, 6),
    ]
    data = [['Startdatum', 'Einddatum', 'Ritnummer', 'Type', 'Afstand (km)', 'Bedrag (€)', 'Status']]
    weekend_rows: list[int] = []
    private_rows: list[int] = []
    total_km = Decimal('0')
    total_amount = Decimal('0')
    private_km = Decimal('0')
    private_amount = Decimal('0')
    weekend_km = Decimal('0')
    weekend_amount = Decimal('0')
    weekday_km = Decimal('0')
    weekday_amount = Decimal('0')
    # Subtotalen per ritnummer, zodat een ritnummerwissel binnen de periode
    # zichtbaar blijft in de uitdraai.
    per_rit: dict[str, list] = {}
    for idx, e in enumerate(events, start=1):
        is_priv = bool(getattr(e, 'is_private', False))
        is_wknd = bool(e.start_at and e.start_at.isoweekday() >= 6)
        rit = (getattr(e, 'ritnummer', '') or '').strip()
        if is_priv:
            type_label = 'Privé'
            status = 'Privé'
            private_rows.append(idx)
            private_km += Decimal(e.distance_km)
            private_amount += Decimal(e.amount)
        else:
            if is_wknd:
                type_label = 'Weekend'
                weekend_rows.append(idx)
                weekend_km += Decimal(e.distance_km)
                weekend_amount += Decimal(e.amount)
            else:
                type_label = 'Doordeweeks'
                weekday_km += Decimal(e.distance_km)
                weekday_amount += Decimal(e.amount)
            status = 'Gefactureerd' if e.invoiced_at else 'Open'
            total_km += Decimal(e.distance_km)
            total_amount += Decimal(e.amount)
            bucket = per_rit.setdefault(rit, [Decimal('0'), Decimal('0'), 0])
            bucket[0] += Decimal(e.distance_km)
            bucket[1] += Decimal(e.amount)
            bucket[2] += 1
        data.append([
            e.start_at.strftime('%Y-%m-%d %H:%M'),
            e.end_at.strftime('%Y-%m-%d %H:%M'),
            rit,
            type_label,
            f'{float(e.distance_km):.3f}',
            f'{float(e.amount):.2f}',
            status,
        ])
    # Subtotal rows
    data.append(['', 'Doordeweeks', '', '', f'{float(weekday_km):.3f}', f'{float(weekday_amount):.2f}', ''])
    data.append(['', 'Weekend', '', '', f'{float(weekend_km):.3f}', f'{float(weekend_amount):.2f}', ''])
    has_private = private_km > 0 or private_amount > 0
    if has_private:
        data.append(['', 'Privé (niet doorbelast)', '', '', f'{float(private_km):.3f}', f'{float(private_amount):.2f}', ''])
    data.append(['', 'Totaal tolkosten', '', '', f'{float(total_km):.3f}', f'{float(total_amount):.2f}', ''])

    tbl = Table(data, repeatRows=1, hAlign='LEFT')
    subtotal_rows = 4 if has_private else 3
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (4, 1), (5, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('BACKGROUND', (0, -subtotal_rows), (-1, -subtotal_rows), colors.HexColor('#eff6ff')),
        ('BACKGROUND', (0, -subtotal_rows + 1), (-1, -subtotal_rows + 1), colors.HexColor('#fef3c7')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f3f4f6')),
        ('FONTNAME', (0, -subtotal_rows), (-1, -1), 'Helvetica-Bold'),
    ]
    if has_private:
        style_cmds.append(('BACKGROUND', (0, -2), (-1, -2), colors.HexColor('#ede9fe')))
        style_cmds.append(('TEXTCOLOR', (0, -2), (-1, -2), colors.HexColor('#5b21b6')))
    for ri in weekend_rows:
        style_cmds.append(('BACKGROUND', (0, ri), (-1, ri), colors.HexColor('#fef9c3')))
    for ri in private_rows:
        style_cmds.append(('BACKGROUND', (0, ri), (-1, ri), colors.HexColor('#ede9fe')))
        style_cmds.append(('TEXTCOLOR', (0, ri), (-1, ri), colors.HexColor('#5b21b6')))
    tbl.setStyle(TableStyle(style_cmds))
    elems.append(tbl)

    # Losse tabel met de totalen per ritnummer wanneer de wagen in deze
    # periode onder meer dan één ritnummer heeft gereden.
    if len(per_rit) > 1:
        elems.append(Spacer(1, 10))
        elems.append(Paragraph('Totalen per ritnummer', styles['Heading3']))
        rit_data = [['Ritnummer', 'Regels', 'Afstand (km)', 'Bedrag (€)']]
        for rit in sorted(per_rit):
            km, bedrag, aantal = per_rit[rit]
            rit_data.append([
                rit or 'Geen ritnummer', str(aantal),
                f'{float(km):.3f}', f'{float(bedrag):.2f}',
            ])
        rit_tbl = Table(rit_data, hAlign='LEFT')
        rit_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ]))
        elems.append(rit_tbl)

    doc.build(elems)
    return buf.getvalue()
