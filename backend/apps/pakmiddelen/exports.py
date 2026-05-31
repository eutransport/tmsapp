"""Pure rendering helpers for pakmiddelen exports.

Returns bytes (no HTTP layer) so they can be reused by the daily report
e-mail as attachments and by the export endpoint as downloads.
"""
from __future__ import annotations

import io

from django.utils import timezone


def build_xlsx_bytes(rows, range_label: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Pakmiddelen'
    headers = ['Datum', 'Ritnummer', 'Pakmiddelen teruggavebon', 'Onderwerp', 'Ontvangen op']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='374151')
        cell.alignment = Alignment(horizontal='left', vertical='center')
    green = PatternFill('solid', fgColor='DCFCE7')
    red = PatternFill('solid', fgColor='FEE2E2')
    for r in rows:
        check_date, ritnummer, has_bon, subject, received = r
        received_str = ''
        if received:
            local_r = timezone.localtime(received) if timezone.is_aware(received) else received
            received_str = local_r.strftime('%d-%m-%Y %H:%M')
        ws.append([
            check_date.strftime('%d-%m-%Y') if check_date else '',
            ritnummer or '',
            'Ja' if has_bon else 'Nee',
            subject or '',
            received_str,
        ])
        row_idx = ws.max_row
        ws.cell(row=row_idx, column=3).fill = green if has_bon else red
    widths = [12, 16, 26, 60, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf_bytes(rows, range_label: str) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1 * cm, rightMargin=1 * cm,
        topMargin=1 * cm, bottomMargin=1 * cm,
        title=f'Pakmiddelen {range_label}',
    )
    styles = getSampleStyleSheet()
    elems = [
        Paragraph(f'<b>Pakmiddelen Teruggavebonnen</b> — {range_label.replace("_", " ")}', styles['Title']),
        Spacer(1, 0.4 * cm),
    ]
    data = [['Datum', 'Ritnummer', 'Bon?', 'Onderwerp', 'Ontvangen op']]
    for r in rows:
        check_date, ritnummer, has_bon, subject, received = r
        received_str = ''
        if received:
            local_r = timezone.localtime(received) if timezone.is_aware(received) else received
            received_str = local_r.strftime('%d-%m-%Y %H:%M')
        data.append([
            check_date.strftime('%d-%m-%Y') if check_date else '',
            ritnummer or '',
            'Ja' if has_bon else 'Nee',
            (subject or '')[:80],
            received_str,
        ])
    if len(data) == 1:
        data.append(['—', '—', '—', 'Geen resultaten in deze periode', '—'])

    table = Table(data, colWidths=[2.5 * cm, 3 * cm, 1.6 * cm, 14 * cm, 4 * cm], repeatRows=1)
    ts = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
    ])
    for i, row in enumerate(data[1:], start=1):
        if row[2] == 'Ja':
            ts.add('BACKGROUND', (2, i), (2, i), colors.HexColor('#DCFCE7'))
            ts.add('TEXTCOLOR', (2, i), (2, i), colors.HexColor('#166534'))
        elif row[2] == 'Nee':
            ts.add('BACKGROUND', (2, i), (2, i), colors.HexColor('#FEE2E2'))
            ts.add('TEXTCOLOR', (2, i), (2, i), colors.HexColor('#991B1B'))
    table.setStyle(ts)
    elems.append(table)
    doc.build(elems)
    return buf.getvalue()
