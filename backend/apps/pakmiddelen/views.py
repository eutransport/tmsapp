"""Pakmiddelen views."""
from __future__ import annotations

import io
import logging
import re
from datetime import datetime

from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.core.permissions import HasModulePermission
from apps.fleet.models import Vehicle

from .imap_service import test_connection
from .graph_service import test_connection_graph
from .models import (
    PakmiddelenAuditLog,
    PakmiddelenCheckResult,
    PakmiddelenConfig,
    PakmiddelenMailLog,
    PakmiddelenRitnummerSelection,
)
from .notifier import send_test_email
from .serializers import (
    PakmiddelenCheckResultSerializer,
    PakmiddelenConfigSerializer,
    PakmiddelenMailLogSerializer,
    PakmiddelenRitnummerSelectionSerializer,
    VehicleRitnummerSerializer,
)
from .services import run_check

logger = logging.getLogger(__name__)


def _client_ip(request):
    xff = request.META.get('HTTP_X_FORWARDED_FOR')
    if xff:
        return xff.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


class _ManagePerm(HasModulePermission):
    """Read => view_pakmiddelen, write => manage_pakmiddelen."""

    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        if request.user.is_superuser or request.user.rol == 'admin':
            return True
        if request.method in ('GET', 'HEAD', 'OPTIONS'):
            return request.user.has_module_permission('view_pakmiddelen')
        return request.user.has_module_permission('manage_pakmiddelen')


class PakmiddelenConfigViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated, _ManagePerm]

    def list(self, request):
        cfg = PakmiddelenConfig.get_solo()
        return Response(PakmiddelenConfigSerializer(cfg).data)

    @action(detail=False, methods=['get', 'put', 'patch'], url_path='current')
    def current(self, request):
        cfg = PakmiddelenConfig.get_solo()
        if request.method == 'GET':
            return Response(PakmiddelenConfigSerializer(cfg).data)
        partial = request.method == 'PATCH'
        ser = PakmiddelenConfigSerializer(cfg, data=request.data, partial=partial, context={'request': request})
        ser.is_valid(raise_exception=True)
        ser.save()
        PakmiddelenAuditLog.objects.create(
            action='update_config', user=request.user, ip_address=_client_ip(request),
            details={'fields': list(request.data.keys())},
        )
        return Response(ser.data)

    @action(detail=False, methods=['post'], url_path='test-imap')
    def test_imap(self, request):
        cfg = PakmiddelenConfig.get_solo()
        # Allow overriding password for the test without persisting it.
        override_pw = (request.data or {}).get('imap_password')
        if override_pw:
            cfg.imap_password = override_pw  # not saved - in-memory only
        host = (request.data or {}).get('imap_host')
        if host:
            cfg.imap_host = host
        result = test_connection(cfg)
        PakmiddelenAuditLog.objects.create(
            action='test_imap', user=request.user, ip_address=_client_ip(request),
            details={'success': result.get('success')},
        )
        return Response(result)

    @action(detail=False, methods=['post'], url_path='test-graph')
    def test_graph(self, request):
        """Test the Microsoft Graph credentials. Optional overrides in body
        (tenant_id, client_id, client_secret, mailbox, folder) are applied
        in-memory only and not persisted."""
        cfg = PakmiddelenConfig.get_solo()
        body = request.data or {}
        for src, dst in [
            ('graph_tenant_id', 'graph_tenant_id'),
            ('graph_client_id', 'graph_client_id'),
            ('graph_client_secret', 'graph_client_secret'),
            ('graph_mailbox', 'graph_mailbox'),
            ('graph_folder', 'graph_folder'),
        ]:
            val = body.get(src)
            if val:
                setattr(cfg, dst, val)
        result = test_connection_graph(cfg)
        PakmiddelenAuditLog.objects.create(
            action='test_graph', user=request.user, ip_address=_client_ip(request),
            details={'success': result.get('success')},
        )
        return Response(result)

    @action(detail=False, methods=['post'], url_path='test-email')
    def test_email(self, request):
        cfg = PakmiddelenConfig.get_solo()
        recipient = (request.data or {}).get('recipient')
        if not recipient:
            return Response({'success': False, 'message': 'Ontvanger is verplicht.'},
                            status=status.HTTP_400_BAD_REQUEST)
        try:
            send_test_email(cfg, recipient, user=request.user)
        except Exception as exc:
            logger.warning('test_email failed: %s', exc)
            PakmiddelenAuditLog.objects.create(
                action='test_email_failed', user=request.user, ip_address=_client_ip(request),
                details={'error': str(exc)},
            )
            return Response({'success': False, 'message': str(exc)})
        PakmiddelenAuditLog.objects.create(
            action='test_email', user=request.user, ip_address=_client_ip(request),
            details={'recipient': recipient},
        )
        return Response({'success': True, 'message': f'Testmail verstuurd naar {recipient}.'})

    @action(detail=False, methods=['post'], url_path='run-now')
    def run_now(self, request):
        cfg = PakmiddelenConfig.get_solo()
        body = request.data or {}
        target_date_iso = body.get('date')
        from_iso = body.get('from')
        to_iso = body.get('to')

        def _p(v):
            if not v:
                return None
            try:
                return datetime.fromisoformat(v).date()
            except ValueError:
                return None

        send_report = bool(body.get('send_report', True))
        df = _p(from_iso)
        dt = _p(to_iso)
        if df and dt:
            if dt < df:
                df, dt = dt, df
            from datetime import timedelta
            cur = df
            totals = {'matched': 0, 'missing_total': 0, 'days': 0, 'errors': []}
            last = None
            while cur <= dt:
                r = run_check(config=cfg, target_date=cur,
                              user=request.user, ip_address=_client_ip(request),
                              send_report=False)
                totals['days'] += 1
                totals['matched'] += int(r.get('matched') or 0)
                totals['missing_total'] += len(r.get('missing') or [])
                if not r.get('success'):
                    totals['errors'].append(f"{cur.isoformat()}: {r.get('message','?')}")
                last = r
                cur = cur + timedelta(days=1)
            msg = (f"{totals['days']} dag(en) gescand: "
                   f"{totals['matched']} match(es), {totals['missing_total']} ontbrekend.")
            if totals['errors']:
                msg += ' Fouten: ' + '; '.join(totals['errors'][:3])
            return Response({
                'success': not totals['errors'],
                'message': msg,
                'matched': totals['matched'],
                'missing_total': totals['missing_total'],
                'days': totals['days'],
                'last': last,
            })

        target_date = _p(target_date_iso)
        if target_date_iso and not target_date:
            return Response({'success': False, 'message': 'Ongeldige datum.'},
                            status=status.HTTP_400_BAD_REQUEST)
        result = run_check(config=cfg, target_date=target_date,
                           user=request.user, ip_address=_client_ip(request),
                           send_report=send_report)
        return Response(result)


class PakmiddelenRitnummerSelectionViewSet(viewsets.ModelViewSet):
    queryset = PakmiddelenRitnummerSelection.objects.all()
    serializer_class = PakmiddelenRitnummerSelectionSerializer
    permission_classes = [IsAuthenticated, _ManagePerm]

    @action(detail=False, methods=['get'], url_path='available-vehicles')
    def available_vehicles(self, request):
        qs = Vehicle.objects.filter(actief=True).exclude(ritnummer='').order_by('ritnummer')
        return Response(VehicleRitnummerSerializer(qs, many=True).data)

    @action(detail=False, methods=['post'], url_path='bulk-set')
    def bulk_set(self, request):
        """
        Replace the active selection set with the given list.
        Body: {"items": [{"ritnummer": "...", "vehicle": "<uuid|null>", "actief": true}, ...]}
        Existing rows for ritnummers not in the list become inactive.
        """
        items = (request.data or {}).get('items')
        if not isinstance(items, list):
            return Response({'detail': 'items moet een lijst zijn.'},
                            status=status.HTTP_400_BAD_REQUEST)
        seen = set()
        for item in items:
            ritnummer = (item.get('ritnummer') or '').strip()
            if not ritnummer:
                continue
            seen.add(ritnummer)
            defaults = {
                'vehicle_id': item.get('vehicle') or None,
                'actief': bool(item.get('actief', True)),
                'notitie': item.get('notitie', '') or '',
            }
            PakmiddelenRitnummerSelection.objects.update_or_create(
                ritnummer=ritnummer, defaults=defaults,
            )
        if seen:
            # Hard-delete rit selections that are no longer in the list,
            # plus their historical check results (so the overview/PDF stop
            # showing removed ritnummers).
            removed = list(
                PakmiddelenRitnummerSelection.objects
                .exclude(ritnummer__in=seen)
                .values_list('ritnummer', flat=True)
            )
            if removed:
                PakmiddelenRitnummerSelection.objects.filter(ritnummer__in=removed).delete()
                PakmiddelenCheckResult.objects.filter(ritnummer__in=removed).delete()
        PakmiddelenAuditLog.objects.create(
            action='bulk_set_ritnummers', user=request.user, ip_address=_client_ip(request),
            details={'count': len(seen)},
        )
        qs = PakmiddelenRitnummerSelection.objects.all()
        return Response(PakmiddelenRitnummerSelectionSerializer(qs, many=True).data)


class PakmiddelenCheckResultViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = PakmiddelenCheckResult.objects.all()
    serializer_class = PakmiddelenCheckResultSerializer
    permission_classes = [IsAuthenticated, _ManagePerm]
    pagination_class = None

    @staticmethod
    def _parse_date(value):
        if not value:
            return None
        try:
            return datetime.fromisoformat(value).date()
        except ValueError:
            return None

    def _filter_qs(self, qs):
        params = self.request.query_params
        date_str = params.get('date')
        from_str = params.get('from')
        to_str = params.get('to')
        d = self._parse_date(date_str)
        df = self._parse_date(from_str)
        dt = self._parse_date(to_str)
        if df and dt:
            if dt < df:
                df, dt = dt, df
            qs = qs.filter(check_date__gte=df, check_date__lte=dt)
        elif d:
            qs = qs.filter(check_date=d)
        elif not (df or dt):
            qs = qs.filter(check_date=timezone.localdate())
        else:
            if df:
                qs = qs.filter(check_date__gte=df)
            if dt:
                qs = qs.filter(check_date__lte=dt)
        return qs

    def get_queryset(self):
        return self._filter_qs(super().get_queryset()).order_by('check_date', 'ritnummer')

    @action(detail=False, methods=['get'], url_path='dates')
    def dates(self, request):
        """Return distinct check_dates and global min/max — used by the UI
        to enable/disable previous/next period buttons."""
        from django.db.models import Min, Max
        qs = PakmiddelenCheckResult.objects.all()
        agg = qs.aggregate(min_d=Min('check_date'), max_d=Max('check_date'))
        params = request.query_params
        df = self._parse_date(params.get('from'))
        dt = self._parse_date(params.get('to'))
        scoped = qs
        if df:
            scoped = scoped.filter(check_date__gte=df)
        if dt:
            scoped = scoped.filter(check_date__lte=dt)
        dates = list(scoped.values_list('check_date', flat=True).distinct().order_by('check_date'))
        return Response({
            'dates': [d.isoformat() for d in dates],
            'min_date': agg['min_d'].isoformat() if agg['min_d'] else None,
            'max_date': agg['max_d'].isoformat() if agg['max_d'] else None,
        })

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        """Export filtered results as xlsx or pdf. Use ?fmt=xlsx|pdf."""
        # Avoid `format` query-param: DRF reserves it for content negotiation
        # (would 404 because no xlsx/pdf renderer is registered).
        fmt = (request.query_params.get('fmt') or 'xlsx').lower()
        qs = self.get_queryset()
        rows = list(qs.values_list(
            'check_date', 'ritnummer', 'has_bon',
            'matched_subject', 'mail_received_at',
        ))
        params = request.query_params
        df = self._parse_date(params.get('from'))
        dt = self._parse_date(params.get('to'))
        single = self._parse_date(params.get('date'))
        if df and dt:
            range_label = f'{df.isoformat()}_tot_{dt.isoformat()}'
        elif single:
            range_label = single.isoformat()
        else:
            range_label = timezone.localdate().isoformat()

        if fmt == 'pdf':
            return self._export_pdf(rows, range_label)
        return self._export_xlsx(rows, range_label)

    def _export_xlsx(self, rows, range_label: str) -> HttpResponse:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = 'Pakmiddelen'
        headers = ['Datum', 'Ritnummer', 'Pakmiddelen teruggavebon', 'Onderwerp', 'Ontvangen op']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='374151')
            cell.alignment = Alignment(horizontal='left', vertical='center')
        green = PatternFill('solid', fgColor='DCFCE7')
        red = PatternFill('solid', fgColor='FEE2E2')
        for r in rows:
            check_date, ritnummer, has_bon, subject, received = r
            received_str = ''
            if received:
                local_r = timezone.localtime(received) if timezone.is_aware(received) else received
                received_str = local_r.strftime('%d-%m-%Y %H:%M')
            ws.append([
                check_date.strftime('%d-%m-%Y') if check_date else '',
                ritnummer or '',
                'Ja' if has_bon else 'Nee',
                subject or '',
                received_str,
            ])
            row_idx = ws.max_row
            ws.cell(row=row_idx, column=3).fill = green if has_bon else red
        # auto-width
        widths = [12, 16, 26, 60, 18]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="pakmiddelen_{range_label}.xlsx"'
        return resp

    def _export_pdf(self, rows, range_label: str) -> HttpResponse:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        )
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            leftMargin=1 * cm, rightMargin=1 * cm,
            topMargin=1 * cm, bottomMargin=1 * cm,
            title=f'Pakmiddelen {range_label}',
        )
        styles = getSampleStyleSheet()
        elems = [
            Paragraph(f'<b>Pakmiddelen Teruggavebonnen</b> — {range_label.replace("_", " ")}', styles['Title']),
            Spacer(1, 0.4 * cm),
        ]
        data = [['Datum', 'Ritnummer', 'Bon?', 'Onderwerp', 'Ontvangen op']]
        for r in rows:
            check_date, ritnummer, has_bon, subject, received = r
            received_str = ''
            if received:
                local_r = timezone.localtime(received) if timezone.is_aware(received) else received
                received_str = local_r.strftime('%d-%m-%Y %H:%M')
            data.append([
                check_date.strftime('%d-%m-%Y') if check_date else '',
                ritnummer or '',
                'Ja' if has_bon else 'Nee',
                (subject or '')[:80],
                received_str,
            ])
        if len(data) == 1:
            data.append(['—', '—', '—', 'Geen resultaten in deze periode', '—'])

        table = Table(data, colWidths=[2.5 * cm, 3 * cm, 1.6 * cm, 14 * cm, 4 * cm], repeatRows=1)
        ts = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ])
        for i, row in enumerate(data[1:], start=1):
            if row[2] == 'Ja':
                ts.add('BACKGROUND', (2, i), (2, i), colors.HexColor('#DCFCE7'))
                ts.add('TEXTCOLOR', (2, i), (2, i), colors.HexColor('#166534'))
            elif row[2] == 'Nee':
                ts.add('BACKGROUND', (2, i), (2, i), colors.HexColor('#FEE2E2'))
                ts.add('TEXTCOLOR', (2, i), (2, i), colors.HexColor('#991B1B'))
        table.setStyle(ts)
        elems.append(table)
        doc.build(elems)
        buf.seek(0)
        resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="pakmiddelen_{range_label}.pdf"'
        return resp

    @action(detail=False, methods=['post'], url_path='mail')
    def mail(self, request):
        """E-mail the current overview to recipients with optional Excel/PDF attachments.

        Body: {
          "recipients": ["x@y"] (optional — falls back to config.notification_recipients),
          "use_config_recipients": bool (default true if no recipients given),
          "include_xlsx": bool, "include_pdf": bool,
          "from": "YYYY-MM-DD", "to": "YYYY-MM-DD"  (optional; defaults to current filter)
        }
        """
        from .notifier import send_overview_report
        body = request.data or {}
        cfg = PakmiddelenConfig.get_solo()

        # Recipients
        extra = body.get('recipients') or []
        if isinstance(extra, str):
            extra = [s.strip() for s in re.split(r'[\s,;]+', extra) if s.strip()]
        use_config = body.get('use_config_recipients', True if not extra else False)
        recipients: list[str] = []
        if use_config:
            recipients.extend(cfg.notification_recipients or [])
        recipients.extend(extra)
        # dedupe preserving order
        seen = set(); recipients = [r for r in recipients if not (r in seen or seen.add(r))]
        if not recipients:
            return Response({'success': False, 'message': 'Geen ontvangers opgegeven.'},
                            status=status.HTTP_400_BAD_REQUEST)

        # Range — fall back to current filter / queryset
        df = self._parse_date(body.get('from') or request.query_params.get('from'))
        dt = self._parse_date(body.get('to') or request.query_params.get('to'))
        single = self._parse_date(body.get('date') or request.query_params.get('date'))
        if df and dt:
            if dt < df:
                df, dt = dt, df
            date_from, date_to = df, dt
        elif single:
            date_from = date_to = single
        else:
            date_from = date_to = timezone.localdate()

        include_xlsx = bool(body.get('include_xlsx', True))
        include_pdf = bool(body.get('include_pdf', False))
        rows = list(
            PakmiddelenCheckResult.objects
            .filter(check_date__gte=date_from, check_date__lte=date_to)
            .order_by('check_date', 'ritnummer')
            .values_list('check_date', 'ritnummer', 'has_bon',
                         'matched_subject', 'mail_received_at')
        )
        range_label = (date_from.isoformat() if date_from == date_to
                       else f'{date_from.isoformat()}_tot_{date_to.isoformat()}')
        xlsx_bytes = self._export_xlsx(rows, range_label).content if include_xlsx else None
        pdf_bytes = self._export_pdf(rows, range_label).content if include_pdf else None

        try:
            send_overview_report(
                cfg, recipients=recipients, date_from=date_from, date_to=date_to,
                xlsx_bytes=xlsx_bytes, pdf_bytes=pdf_bytes, user=request.user,
            )
        except Exception as exc:  # noqa: BLE001
            logger.exception('mail_overview failed: %s', exc)
            return Response({'success': False, 'message': f'Versturen mislukt: {exc}'},
                            status=status.HTTP_200_OK)

        return Response({
            'success': True,
            'message': f'Overzicht verstuurd naar {len(recipients)} ontvanger(s).',
            'recipients': recipients,
        })


class PakmiddelenMailLogViewSet(viewsets.ReadOnlyModelViewSet):
    """Read-only paginated history of all e-mails sent by the pakmiddelen module."""

    queryset = PakmiddelenMailLog.objects.all().select_related('user')
    serializer_class = PakmiddelenMailLogSerializer
    permission_classes = [IsAuthenticated, _ManagePerm]
    # Uses the default SafePageNumberPagination (page_size=25, ?page=&page_size=)

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        mail_type = params.get('mail_type')
        success = params.get('success')
        from_str = params.get('from')
        to_str = params.get('to')
        if mail_type:
            qs = qs.filter(mail_type=mail_type)
        if success in ('true', '1', 'yes'):
            qs = qs.filter(success=True)
        elif success in ('false', '0', 'no'):
            qs = qs.filter(success=False)
        if from_str:
            try:
                qs = qs.filter(sent_at__date__gte=datetime.fromisoformat(from_str).date())
            except ValueError:
                pass
        if to_str:
            try:
                qs = qs.filter(sent_at__date__lte=datetime.fromisoformat(to_str).date())
            except ValueError:
                pass
        return qs.order_by('-sent_at')
